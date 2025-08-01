"""Excel-based BOM calculation service using file-based integration."""

import os
import tempfile
import shutil
from decimal import Decimal
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from pathlib import Path
# import pandas as pd  # Optional - only needed for advanced data processing
import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import (
    SalesOrder, SalesOrderLine, BillOfMaterials, BOMLineItem,
    BOMStatus, ComponentType, Product
)


class ExcelBOMService:
    """Service for calculating BOMs using Excel templates and formulas."""
    
    def __init__(self, db_session: Session, templates_path: str = "excel_templates"):
        self.db = db_session
        self.templates_path = Path(templates_path)
        self.temp_dir = Path(tempfile.gettempdir()) / "bom_calculations"
        self.temp_dir.mkdir(exist_ok=True)
    
    # === Main BOM Calculation Flow ===
    
    def calculate_bom_for_order(self, sales_order_id: int, 
                               template_name: str = "default_bom_template.xlsx",
                               user_id: str = None) -> BillOfMaterials:
        """Calculate BOM for a sales order using Excel template."""
        
        # Get or create BOM record
        bom = self._get_or_create_bom(sales_order_id)
        
        try:
            bom.status = BOMStatus.CALCULATING
            bom.calculation_started = datetime.utcnow()
            bom.calculated_by = user_id
            bom.excel_template_used = template_name
            self.db.commit()
            
            # Load sales order data
            sales_order = self.db.query(SalesOrder).get(sales_order_id)
            if not sales_order:
                raise ValueError(f"Sales order {sales_order_id} not found")
            
            # Prepare input data for Excel
            input_data = self._prepare_input_data(sales_order)
            
            # Process Excel calculation
            excel_results = self._process_excel_calculation(template_name, input_data, bom.bom_number)
            
            # Parse and save results
            self._save_bom_results(bom, excel_results)
            
            # Mark as completed
            bom.status = BOMStatus.COMPLETED
            bom.calculation_completed = datetime.utcnow()
            bom.calculation_duration_seconds = int(
                (bom.calculation_completed - bom.calculation_started).total_seconds()
            )
            
            # Update sales order
            sales_order.bom_calculated = True
            sales_order.bom_calculation_date = bom.calculation_completed
            
            self.db.commit()
            
        except Exception as e:
            bom.status = BOMStatus.ERROR
            bom.excel_calculation_errors = str(e)
            self.db.commit()
            raise
        
        return bom
    
    # === Excel Processing ===
    
    def _process_excel_calculation(self, template_name: str, input_data: Dict[str, Any], 
                                  bom_number: str) -> Dict[str, Any]:
        """Process BOM calculation using Excel template."""
        
        template_path = self.templates_path / template_name
        if not template_path.exists():
            raise FileNotFoundError(f"Excel template not found: {template_path}")
        
        # Create working copy of template
        work_file = self.temp_dir / f"{bom_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(template_path, work_file)
        
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(work_file, data_only=False)
            
            # Populate input data
            self._populate_excel_inputs(workbook, input_data)
            
            # Save to trigger formula calculations
            workbook.save(work_file)
            
            # Reload with calculated values
            workbook = openpyxl.load_workbook(work_file, data_only=True)
            
            # Extract results
            results = self._extract_excel_results(workbook)
            results['excel_file_path'] = str(work_file)
            
            return results
            
        except Exception as e:
            # Clean up on error
            if work_file.exists():
                work_file.unlink()
            raise e
    
    def _populate_excel_inputs(self, workbook: openpyxl.Workbook, input_data: Dict[str, Any]):
        """Populate Excel template with sales order data."""
        
        # Assume template has an "Inputs" sheet
        if "Inputs" not in workbook.sheetnames:
            raise ValueError("Excel template must have an 'Inputs' sheet")
        
        input_sheet = workbook["Inputs"]
        
        # Map input data to specific cells (this would be template-specific)
        cell_mapping = {
            'order_number': 'B2',
            'customer_id': 'B3',
            'order_date': 'B4',
            'total_quantity': 'B5',
        }
        
        # Populate basic order info
        for field, cell in cell_mapping.items():
            if field in input_data:
                input_sheet[cell] = input_data[field]
        
        # Populate line items starting at row 8
        start_row = 8
        for i, line_item in enumerate(input_data.get('line_items', [])):
            row = start_row + i
            input_sheet[f'A{row}'] = line_item['line_number']
            input_sheet[f'B{row}'] = line_item['product_sku']
            input_sheet[f'C{row}'] = line_item['product_name']
            input_sheet[f'D{row}'] = line_item['quantity']
            input_sheet[f'E{row}'] = float(line_item['unit_price'])
    
    def _extract_excel_results(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Extract BOM results from calculated Excel file."""
        
        # Assume template has a "BOM_Results" sheet
        if "BOM_Results" not in workbook.sheetnames:
            raise ValueError("Excel template must have a 'BOM_Results' sheet")
        
        results_sheet = workbook["BOM_Results"]
        
        # Extract summary totals
        results = {
            'total_material_cost': self._get_decimal_cell_value(results_sheet, 'B2'),
            'total_labor_cost': self._get_decimal_cell_value(results_sheet, 'B3'),
            'total_overhead_cost': self._get_decimal_cell_value(results_sheet, 'B4'),
            'total_bom_cost': self._get_decimal_cell_value(results_sheet, 'B5'),
            'line_items': []
        }
        
        # Extract line items (starting from row 8)
        row = 8
        while True:
            line_number = results_sheet[f'A{row}'].value
            if not line_number:
                break
            
            component_name = results_sheet[f'B{row}'].value
            if not component_name:
                break
            
            line_item = {
                'line_number': int(line_number),
                'component_sku': results_sheet[f'C{row}'].value,
                'component_name': component_name,
                'component_description': results_sheet[f'D{row}'].value,
                'component_type': results_sheet[f'E{row}'].value or 'raw_material',
                'quantity_required': self._get_decimal_cell_value(results_sheet, f'F{row}'),
                'unit_of_measure': results_sheet[f'G{row}'].value or 'EA',
                'unit_cost': self._get_decimal_cell_value(results_sheet, f'H{row}'),
                'extended_cost': self._get_decimal_cell_value(results_sheet, f'I{row}'),
                'excel_row_reference': f'Row {row}',
                'cost_source': 'excel'
            }
            
            results['line_items'].append(line_item)
            row += 1
        
        return results
    
    def _get_decimal_cell_value(self, sheet, cell_ref: str) -> Decimal:
        """Safely extract decimal value from Excel cell."""
        value = sheet[cell_ref].value
        if value is None:
            return Decimal('0')
        return Decimal(str(value))
    
    # === Data Preparation ===
    
    def _prepare_input_data(self, sales_order: SalesOrder) -> Dict[str, Any]:
        """Prepare sales order data for Excel input."""
        
        line_items = []
        for line in sales_order.lines:
            line_items.append({
                'line_number': line.line_number,
                'product_sku': line.product_sku,
                'product_name': line.product_name,
                'quantity': line.quantity_ordered,
                'unit_price': line.unit_price,
                'extended_price': line.extended_price
            })
        
        return {
            'order_number': sales_order.order_number,
            'customer_id': sales_order.customer_id,
            'customer_name': sales_order.customer_name,
            'order_date': sales_order.order_date,
            'total_quantity': sum(line.quantity_ordered for line in sales_order.lines),
            'subtotal': sales_order.subtotal,
            'line_items': line_items
        }
    
    def _save_bom_results(self, bom: BillOfMaterials, excel_results: Dict[str, Any]):
        """Save Excel calculation results to BOM records."""
        
        # Update BOM header totals
        bom.total_material_cost = excel_results['total_material_cost']
        bom.total_labor_cost = excel_results['total_labor_cost']
        bom.total_overhead_cost = excel_results['total_overhead_cost']
        bom.total_bom_cost = excel_results['total_bom_cost']
        bom.excel_file_path = excel_results.get('excel_file_path')
        
        # Clear existing line items
        self.db.query(BOMLineItem).filter(BOMLineItem.bom_id == bom.id).delete()
        
        # Create new line items
        for item_data in excel_results['line_items']:
            # Try to match component to existing product
            product_id = None
            if item_data.get('component_sku'):
                product = self.db.query(Product).filter(
                    Product.sku == item_data['component_sku']
                ).first()
                if product:
                    product_id = product.id
            
            # Map component type
            component_type = self._map_component_type(item_data.get('component_type', 'raw_material'))
            
            line_item = BOMLineItem(
                bom_id=bom.id,
                line_number=item_data['line_number'],
                component_sku=item_data.get('component_sku'),
                component_name=item_data['component_name'],
                component_description=item_data.get('component_description'),
                component_type=component_type,
                quantity_required=item_data['quantity_required'],
                unit_of_measure=item_data['unit_of_measure'],
                unit_cost=item_data['unit_cost'],
                extended_cost=item_data['extended_cost'],
                excel_row_reference=item_data.get('excel_row_reference'),
                cost_source=item_data.get('cost_source', 'excel'),
                product_id=product_id,
                requires_procurement=product_id is not None
            )
            
            line_item.calculate_extended_cost()  # Ensure calculation is correct
            self.db.add(line_item)
        
        # Recalculate totals to ensure consistency
        bom.calculate_totals()
    
    def _map_component_type(self, type_string: str) -> ComponentType:
        """Map Excel component type string to enum."""
        type_mapping = {
            'raw_material': ComponentType.RAW_MATERIAL,
            'component': ComponentType.COMPONENT,
            'assembly': ComponentType.ASSEMBLY,
            'packaging': ComponentType.PACKAGING,
            'labor': ComponentType.LABOR,
            'overhead': ComponentType.OVERHEAD
        }
        
        return type_mapping.get(type_string.lower(), ComponentType.RAW_MATERIAL)
    
    # === BOM Management ===
    
    def _get_or_create_bom(self, sales_order_id: int) -> BillOfMaterials:
        """Get existing BOM or create new one for sales order."""
        
        bom = self.db.query(BillOfMaterials).filter(
            BillOfMaterials.sales_order_id == sales_order_id
        ).first()
        
        if not bom:
            # Generate BOM number
            bom_number = self._generate_bom_number(sales_order_id)
            
            bom = BillOfMaterials(
                sales_order_id=sales_order_id,
                bom_number=bom_number,
                status=BOMStatus.PENDING
            )
            self.db.add(bom)
            self.db.flush()  # Get ID
        
        return bom
    
    def _generate_bom_number(self, sales_order_id: int) -> str:
        """Generate unique BOM number."""
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        return f"BOM-{sales_order_id}-{timestamp}"
    
    # === Template Management ===
    
    def create_bom_template(self, template_name: str) -> Path:
        """Create a basic BOM calculation template."""
        
        template_path = self.templates_path / template_name
        template_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook with required sheets
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create Inputs sheet
        inputs_sheet = workbook.create_sheet("Inputs")
        self._setup_inputs_sheet(inputs_sheet)
        
        # Create BOM_Results sheet
        results_sheet = workbook.create_sheet("BOM_Results")
        self._setup_results_sheet(results_sheet)
        
        # Create Calculations sheet (for formulas)
        calc_sheet = workbook.create_sheet("Calculations")
        self._setup_calculations_sheet(calc_sheet)
        
        # Save template
        workbook.save(template_path)
        return template_path
    
    def _setup_inputs_sheet(self, sheet):
        """Set up the inputs sheet structure."""
        sheet['A1'] = "Sales Order Inputs"
        sheet['A2'] = "Order Number:"
        sheet['A3'] = "Customer ID:"
        sheet['A4'] = "Order Date:"
        sheet['A5'] = "Total Quantity:"
        
        # Line items header
        sheet['A7'] = "Line Items:"
        sheet['A8'] = "Line #"
        sheet['B8'] = "SKU"
        sheet['C8'] = "Product Name"
        sheet['D8'] = "Quantity"
        sheet['E8'] = "Unit Price"
    
    def _setup_results_sheet(self, sheet):
        """Set up the BOM results sheet structure."""
        sheet['A1'] = "BOM Calculation Results"
        sheet['A2'] = "Total Material Cost:"
        sheet['A3'] = "Total Labor Cost:"
        sheet['A4'] = "Total Overhead Cost:"
        sheet['A5'] = "Total BOM Cost:"
        
        # BOM line items header
        sheet['A7'] = "BOM Components:"
        sheet['A8'] = "Line #"
        sheet['B8'] = "Component Name"
        sheet['C8'] = "Component SKU"
        sheet['D8'] = "Description"
        sheet['E8'] = "Type"
        sheet['F8'] = "Quantity"
        sheet['G8'] = "UOM"
        sheet['H8'] = "Unit Cost"
        sheet['I8'] = "Extended Cost"
    
    def _setup_calculations_sheet(self, sheet):
        """Set up the calculations sheet with sample formulas."""
        sheet['A1'] = "BOM Calculation Formulas"
        sheet['A3'] = "Sample calculation logic would go here"
        sheet['A4'] = "This sheet contains the business logic for:"
        sheet['A5'] = "- Material requirements calculation"
        sheet['A6'] = "- Labor time estimation"
        sheet['A7'] = "- Overhead allocation"
        sheet['A8'] = "- Cost rollup calculations"
    
    # === Utility Methods ===
    
    def get_bom_status(self, sales_order_id: int) -> Optional[BOMStatus]:
        """Get current BOM calculation status for an order."""
        bom = self.db.query(BillOfMaterials).filter(
            BillOfMaterials.sales_order_id == sales_order_id
        ).first()
        
        return bom.status if bom else None
    
    def cleanup_temp_files(self, older_than_hours: int = 24):
        """Clean up temporary Excel files older than specified hours."""
        cutoff_time = datetime.now().timestamp() - (older_than_hours * 3600)
        
        for file_path in self.temp_dir.glob("*.xlsx"):
            if file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()
                except OSError:
                    pass  # File might be in use